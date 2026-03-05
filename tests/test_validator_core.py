"""
tests/test_validator_core.py
============================
Tests unitarios para la lógica core del validador.
Ejecutar con: pytest -q

Lógica de semáforo por corte (v2):
  - error < REPRO              → VERDE
  - REPRO ≤ error < ADMISIBLE → AMARILLO
  - error ≥ ADMISIBLE          → ROJO
"""
from __future__ import annotations

import io
import pytest
import pandas as pd
import numpy as np

from core.models import ThresholdConfig, ValidationResult
from core.validator_core import (
    # Normalización
    strip_accents,
    canon_prop,
    canon_corte,
    # Emparejamiento
    _nombre_base_crudo,
    pair_files,
    # Lectura
    read_file,
    # Reglas de evaluación
    es_corte_pesado,
    _buscar_umbrales,
    _prop_base_para_umbral,
    clasificar_propiedad,
    # Umbrales
    detectar_columna_tipo,
    normalizar_tipo,
    construir_umbrales,
    crear_semantica_alias,
    # Pipeline
    calcular_errores_crudo_df,
    _sem_global_por_crudo,
    _build_summary_df,
    validate_params,
    build_excel,
    run_validation,
    # Alias compat
    canonize_name,
    validate_thresholds,
    # Constantes
    SEMAFORO_COLORS,
    SEMAFORO_LABELS,
    SUPPORTED_EXTENSIONS,
    DEFAULT_TOL,
    DEFAULT_TOL_PESADOS,
    DEFAULT_PCT_OK_AMARILLO,
    DEFAULT_PCT_ROJO_ROJO,
)


# ===========================================================================
# Fixtures
# ===========================================================================

@pytest.fixture
def alias_prop():
    return crear_semantica_alias()


@pytest.fixture
def simple_isa() -> pd.DataFrame:
    return pd.DataFrame({
        "Propiedad": ["Densidad", "Viscosidad", "Azufre"],
        "150-200": [850.0, 5.0, 0.10],
        "200-250": [860.0, 6.0, 0.20],
        "300-350": [870.0, 7.0, 0.35],
    })


@pytest.fixture
def simple_rams() -> pd.DataFrame:
    return pd.DataFrame({
        "Propiedad": ["Densidad", "Viscosidad", "Azufre"],
        "150-200": [851.0, 5.5, 0.11],
        "200-250": [858.5, 6.0, 0.22],
        "300-350": [872.0, 6.5, 0.36],
    })


@pytest.fixture
def simple_umbrales() -> dict:
    """
    Umbrales en el nuevo formato: {(prop, corte): (repro, admisible)}.

    Densidad: REPRO=2.0, ADMISIBLE=4.0
    Viscosidad 50: REPRO=1.0, ADMISIBLE=2.0
    Azufre: REPRO=0.05, ADMISIBLE=0.10
    """
    return {
        ("DENSIDAD",     "150-200"): (2.0, 4.0),
        ("DENSIDAD",     "200-250"): (2.0, 4.0),
        ("DENSIDAD",     "300-350"): (2.0, 4.0),
        ("VISCOSIDAD 50","150-200"): (1.0, 2.0),
        ("VISCOSIDAD 50","200-250"): (1.0, 2.0),
        ("VISCOSIDAD 50","300-350"): (1.0, 2.0),
        ("AZUFRE",       "150-200"): (0.05, 0.10),
        ("AZUFRE",       "200-250"): (0.05, 0.10),
        ("AZUFRE",       "300-350"): (0.05, 0.10),
    }


@pytest.fixture
def isa_bytes(simple_isa) -> bytes:
    buf = io.BytesIO()
    simple_isa.to_excel(buf, index=False)
    return buf.getvalue()


@pytest.fixture
def rams_bytes(simple_rams) -> bytes:
    buf = io.BytesIO()
    simple_rams.to_excel(buf, index=False)
    return buf.getvalue()


@pytest.fixture
def matriz_bytes() -> bytes:
    """
    Matriz de umbrales con filas REPRO y ADMISIBLE para DENSIDAD, VISCOSIDAD 50 y AZUFRE.
    """
    rows = [
        # DENSIDAD
        {"Propiedad": "DENSIDAD",      "Tipo": "Reproductibilidad",
         "150-200": 2.0, "200-250": 2.0, "300-350": 2.0},
        {"Propiedad": "DENSIDAD",      "Tipo": "Admisible",
         "150-200": 4.0, "200-250": 4.0, "300-350": 4.0},
        # VISCOSIDAD 50
        {"Propiedad": "VISCOSIDAD 50", "Tipo": "Reproductibilidad",
         "150-200": 1.0, "200-250": 1.0, "300-350": 1.0},
        {"Propiedad": "VISCOSIDAD 50", "Tipo": "Admisible",
         "150-200": 2.0, "200-250": 2.0, "300-350": 2.0},
        # AZUFRE
        {"Propiedad": "AZUFRE",        "Tipo": "Reproductibilidad",
         "150-200": 0.05, "200-250": 0.05, "300-350": 0.05},
        {"Propiedad": "AZUFRE",        "Tipo": "Admisible",
         "150-200": 0.10, "200-250": 0.10, "300-350": 0.10},
    ]
    df = pd.DataFrame(rows)
    buf = io.BytesIO()
    df.to_excel(buf, index=False)
    return buf.getvalue()


# ===========================================================================
# 1. Tests strip_accents
# ===========================================================================

class TestStripAccents:

    def test_removes_accents(self):
        assert strip_accents("Densidád") == "Densidad"

    def test_none_returns_empty(self):
        assert strip_accents(None) == ""

    def test_no_accents_unchanged(self):
        assert strip_accents("DENSIDAD") == "DENSIDAD"

    def test_multiple_accents(self):
        result = strip_accents("Núméro")
        assert result == "Numero"


# ===========================================================================
# 2. Tests canon_prop
# ===========================================================================

class TestCanonProp:

    def test_basic(self):
        assert canon_prop("Densidad") == "DENSIDAD"

    def test_removes_degree(self):
        assert canon_prop("Densidad a 15°C") == "DENSIDAD A 15C"

    def test_alias_applied(self):
        alias = crear_semantica_alias()
        assert canon_prop("NOR CLARO", alias) == "RON"

    def test_alias_carbono_conradson(self):
        alias = crear_semantica_alias()
        assert canon_prop("Carbono Conradson", alias) == "RESIDUO DE CARBON"

    def test_none_returns_empty(self):
        assert canon_prop(None) == ""

    def test_only_symbols_returns_empty(self):
        assert canon_prop("---") == ""


# ===========================================================================
# 3. Tests canon_corte
# ===========================================================================

class TestCanonCorte:

    def test_basic_range(self):
        assert canon_corte("150-200") == "150-200"

    def test_typographic_dash(self):
        assert canon_corte("150\u2013200") == "150-200"

    def test_spaces_around_dash(self):
        assert canon_corte("150 - 200") == "150-200"

    def test_removes_degrees(self):
        assert canon_corte("150°-200°") == "150-200"

    def test_plus_corte(self):
        assert canon_corte("538+") == "538+"

    def test_none_returns_empty(self):
        assert canon_corte(None) == ""


# ===========================================================================
# 4. Tests es_corte_pesado
# ===========================================================================

class TestEsCorteePesado:

    @pytest.mark.parametrize("corte,expected", [
        ("C6-C10", True),
        ("C7", True),
        ("538+", True),
        ("300-350", True),
        ("298-350", False),
        ("150-200", False),
        ("IBP-150", False),
    ])
    def test_parametrized(self, corte, expected):
        assert es_corte_pesado(corte) == expected

    def test_exact_boundary_299(self):
        assert es_corte_pesado("299-350") is True

    def test_exact_boundary_298(self):
        assert es_corte_pesado("298-350") is False


# ===========================================================================
# 5. Tests nombre_base_crudo
# ===========================================================================

class TestNombreBaseCrudo:

    @pytest.mark.parametrize("fname,expected", [
        ("ISA_Crudo_Maya.xlsx", "Crudo_Maya"),
        ("RAMS_Crudo_Maya.xlsx", "Crudo_Maya"),
        ("Crudo_Maya_ISA.xlsx", "Crudo_Maya"),
        ("GMX-2023-1_ISA_v1.xlsx", "GMX-2023-1"),
        ("RAMS_GMX-2023-1.xlsx", "GMX-2023-1"),
        ("COL-2023-1.xlsx", "COL-2023-1"),
    ])
    def test_parametrized(self, fname, expected):
        assert _nombre_base_crudo(fname) == expected

    def test_strips_isa_prefix(self):
        assert "isa" not in _nombre_base_crudo("ISA_Maya.xlsx").lower()

    def test_strips_rams_prefix(self):
        assert "rams" not in _nombre_base_crudo("RAMS_Maya.xlsx").lower()


# ===========================================================================
# 6. Tests pair_files
# ===========================================================================

class TestPairFiles:

    def test_pair_basic(self):
        isa  = ["ISA_Maya.xlsx", "ISA_Brent.xlsx"]
        rams = ["RAMS_Maya.xlsx", "RAMS_Brent.xlsx"]
        paired, u_isa, u_rams = pair_files(isa, rams)
        assert len(paired) == 2
        assert not u_isa
        assert not u_rams

    def test_unpaired_isa(self):
        isa  = ["ISA_Maya.xlsx", "ISA_Extra.xlsx"]
        rams = ["RAMS_Maya.xlsx"]
        _, u_isa, _ = pair_files(isa, rams)
        assert "ISA_Extra.xlsx" in u_isa

    def test_unpaired_rams(self):
        isa  = ["ISA_Maya.xlsx"]
        rams = ["RAMS_Maya.xlsx", "RAMS_Extra.xlsx"]
        _, _, u_rams = pair_files(isa, rams)
        assert "RAMS_Extra.xlsx" in u_rams

    def test_structured_code(self):
        isa  = ["GMX-2023-1_ISA.xlsx"]
        rams = ["GMX-2023-1_RAMS.xlsx"]
        paired, u_isa, u_rams = pair_files(isa, rams)
        assert len(paired) == 1
        assert not u_isa and not u_rams

    def test_empty(self):
        paired, u_isa, u_rams = pair_files([], [])
        assert paired == {}


# ===========================================================================
# 7. Tests read_file
# ===========================================================================

class TestReadFile:

    def test_read_xlsx(self, simple_isa, isa_bytes):
        buf = io.BytesIO(isa_bytes)
        df  = read_file(buf, "test.xlsx")
        assert df.shape == simple_isa.shape

    def test_read_csv_semicolon(self, simple_isa):
        csv = simple_isa.to_csv(index=False, sep=";", decimal=",").encode()
        df  = read_file(io.BytesIO(csv), "test.csv")
        assert df.shape[0] == 3

    def test_unsupported_extension(self):
        with pytest.raises(ValueError, match="soportado"):
            read_file(io.BytesIO(b"data"), "file.xlsm")

    def test_corrupt_xlsx(self):
        with pytest.raises(ValueError):
            read_file(io.BytesIO(b"not an excel"), "corrupt.xlsx")


# ===========================================================================
# 8. Tests construir_umbrales — nuevo formato (repro, admisible)
# ===========================================================================

class TestConstruirUmbrales:

    def test_stores_repro_and_admisible_separately(self, alias_prop):
        df = pd.DataFrame({
            "Propiedad": ["DENSIDAD", "DENSIDAD"],
            "Tipo":      ["Reproductibilidad", "Admisible"],
            "150-200":   [2.0, 4.0],
        })
        umbrales = construir_umbrales(df, alias_prop)
        repro, admis = umbrales[("DENSIDAD", "150-200")]
        assert repro == 2.0
        assert admis == 4.0

    def test_only_repro(self, alias_prop):
        df = pd.DataFrame({
            "Propiedad": ["DENSIDAD"],
            "Tipo":      ["Reproductibilidad"],
            "150-200":   [2.0],
        })
        umbrales = construir_umbrales(df, alias_prop)
        repro, admis = umbrales[("DENSIDAD", "150-200")]
        assert repro == 2.0
        assert admis is None

    def test_only_admisible(self, alias_prop):
        df = pd.DataFrame({
            "Propiedad": ["DENSIDAD"],
            "Tipo":      ["Admisible"],
            "150-200":   [4.0],
        })
        umbrales = construir_umbrales(df, alias_prop)
        repro, admis = umbrales[("DENSIDAD", "150-200")]
        assert repro is None
        assert admis == 4.0

    def test_no_propiedad_col_raises(self, alias_prop):
        df = pd.DataFrame({"Tipo": ["Repro"], "150-200": [1.0]})
        with pytest.raises(ValueError, match="'Propiedad'"):
            construir_umbrales(df, alias_prop)

    def test_no_tipo_col_raises(self, alias_prop):
        df = pd.DataFrame({"Propiedad": ["DENSIDAD"], "150-200": [1.0]})
        with pytest.raises(ValueError, match="Tipo"):
            construir_umbrales(df, alias_prop)

    def test_max_wins_within_type(self, alias_prop):
        """Si hay dos filas REPRO para la misma clave, gana el mayor."""
        df = pd.DataFrame({
            "Propiedad": ["AZUFRE", "AZUFRE"],
            "Tipo":      ["Reproductibilidad", "Reproductibilidad"],
            "150-200":   [0.05, 0.08],
        })
        umbrales = construir_umbrales(df, alias_prop)
        repro, _ = umbrales[("AZUFRE", "150-200")]
        assert repro == 0.08

    def test_non_repro_non_admis_ignored(self, alias_prop):
        df = pd.DataFrame({
            "Propiedad": ["DENSIDAD", "DENSIDAD"],
            "Tipo":      ["Reproductibilidad", "Comentario"],
            "150-200":   [2.0, 99.0],
        })
        umbrales = construir_umbrales(df, alias_prop)
        repro, admis = umbrales[("DENSIDAD", "150-200")]
        assert repro == 2.0
        assert admis is None   # "Comentario" no se procesa


# ===========================================================================
# 9. Tests clasificar_propiedad — nueva lógica REPRO / ADMISIBLE
# ===========================================================================

class TestClasificarPropiedad:

    @pytest.fixture
    def basic_umbrales(self):
        """
        DENSIDAD:      REPRO=2.0, ADMISIBLE=4.0
        VISCOSIDAD 50: REPRO=1.0, ADMISIBLE=2.0
        AZUFRE:        REPRO=0.05, ADMISIBLE=0.10
        PESO:          REPRO=1.0, ADMISIBLE=2.0
        """
        return {
            ("DENSIDAD",     "150-200"): (2.0,  4.0),
            ("VISCOSIDAD 50","150-200"): (1.0,  2.0),
            ("AZUFRE",       "150-200"): (0.05, 0.10),
            ("PESO",         "150-200"): (1.0,  2.0),
        }

    # ── Verde: error < REPRO ──────────────────────────────────────────────

    def test_verde_error_menor_que_repro(self, basic_umbrales):
        errores = {"150-200": 1.5}   # 1.5 < 2.0 (REPRO) → VERDE
        sem, _, _, _, _, _ = clasificar_propiedad(
            errores, "DENSIDAD", basic_umbrales, 0.9, 0.3
        )
        assert sem == "VERDE"

    def test_verde_error_cero(self, basic_umbrales):
        errores = {"150-200": 0.0}
        sem, _, _, _, _, _ = clasificar_propiedad(
            errores, "DENSIDAD", basic_umbrales, 0.9, 0.3
        )
        assert sem == "VERDE"

    # ── Amarillo: REPRO ≤ error < ADMISIBLE ──────────────────────────────

    def test_amarillo_en_repro(self, basic_umbrales):
        errores = {"150-200": 2.0}   # error == REPRO → AMARILLO
        sem, _, _, _, _, _ = clasificar_propiedad(
            errores, "DENSIDAD", basic_umbrales, 0.9, 0.3
        )
        assert sem == "AMARILLO"

    def test_amarillo_entre_repro_y_admisible(self, basic_umbrales):
        errores = {"150-200": 3.0}   # 2.0 ≤ 3.0 < 4.0 → AMARILLO
        sem, _, _, _, _, _ = clasificar_propiedad(
            errores, "DENSIDAD", basic_umbrales, 0.9, 0.3
        )
        assert sem == "AMARILLO"

    def test_amarillo_justo_antes_de_admisible(self, basic_umbrales):
        errores = {"150-200": 3.99}  # 3.99 < 4.0 (ADMISIBLE) → AMARILLO
        sem, _, _, _, _, _ = clasificar_propiedad(
            errores, "DENSIDAD", basic_umbrales, 0.9, 0.3
        )
        assert sem == "AMARILLO"

    # ── Rojo: error ≥ ADMISIBLE ───────────────────────────────────────────

    def test_rojo_en_admisible(self, basic_umbrales):
        errores = {"150-200": 4.0}   # error == ADMISIBLE → ROJO
        sem, _, _, _, _, _ = clasificar_propiedad(
            errores, "DENSIDAD", basic_umbrales, 0.9, 0.3
        )
        assert sem == "ROJO"

    def test_rojo_mayor_que_admisible(self, basic_umbrales):
        errores = {"150-200": 99.0}
        sem, _, _, _, _, _ = clasificar_propiedad(
            errores, "DENSIDAD", basic_umbrales, 0.9, 0.3
        )
        assert sem == "ROJO"

    # ── Sin ADMISIBLE: solo REPRO como frontera ───────────────────────────

    def test_sin_admisible_verde(self):
        umbrales = {("DENSIDAD", "150-200"): (2.0, None)}
        errores  = {"150-200": 1.5}   # < 2.0 → VERDE
        sem, _, _, _, _, _ = clasificar_propiedad(
            errores, "DENSIDAD", umbrales, 0.9, 0.3
        )
        assert sem == "VERDE"

    def test_sin_admisible_rojo(self):
        umbrales = {("DENSIDAD", "150-200"): (2.0, None)}
        errores  = {"150-200": 3.0}   # ≥ 2.0, sin ADMISIBLE → ROJO
        sem, _, _, _, _, _ = clasificar_propiedad(
            errores, "DENSIDAD", umbrales, 0.9, 0.3
        )
        assert sem == "ROJO"

    # ── Sin REPRO: solo ADMISIBLE como frontera ───────────────────────────

    def test_sin_repro_amarillo(self):
        """Sin REPRO, no se puede garantizar VERDE: error < ADMISIBLE → AMARILLO."""
        umbrales = {("DENSIDAD", "150-200"): (None, 4.0)}
        errores  = {"150-200": 2.0}   # No hay REPRO; 2.0 < 4.0 (ADMISIBLE) → AMARILLO
        sem, estados, _, _, _, _ = clasificar_propiedad(
            errores, "DENSIDAD", umbrales, 0.9, 0.3
        )
        assert estados["150-200"] == "AMARILLO"
        assert sem == "AMARILLO"

    def test_sin_repro_rojo(self):
        umbrales = {("DENSIDAD", "150-200"): (None, 4.0)}
        errores  = {"150-200": 5.0}   # ≥ 4.0 (ADMISIBLE) → ROJO
        sem, estados, _, _, _, _ = clasificar_propiedad(
            errores, "DENSIDAD", umbrales, 0.9, 0.3
        )
        assert estados["150-200"] == "ROJO"
        assert sem == "ROJO"

    # ── NA: sin umbral ────────────────────────────────────────────────────

    def test_na_sin_umbral(self):
        errores = {"150-200": 1.0}
        sem, _, _, _, _, _ = clasificar_propiedad(
            errores, "SIN_UMBRAL", {}, 0.9, 0.3
        )
        assert sem == "NA"

    # ── Sin valores numéricos ─────────────────────────────────────────────

    def test_no_numerico_skipped(self, basic_umbrales):
        errores = {"150-200": None, "200-250": None}
        sem, estados, _, _, _, _ = clasificar_propiedad(
            errores, "DENSIDAD", basic_umbrales, 0.9, 0.3
        )
        assert sem == ""
        assert all(e == "(no numérico)" for e in estados.values())

    # ── Fallback PESO ACUMULADO → PESO ────────────────────────────────────

    def test_peso_acumulado_fallback(self):
        umbrales = {("PESO", "150-200"): (1.0, 2.0)}
        errores  = {"150-200": 0.5}   # < 1.0 (REPRO) → VERDE
        sem, _, _, _, _, _ = clasificar_propiedad(
            errores, "PESO ACUMULADO", umbrales, 0.9, 0.3
        )
        assert sem == "VERDE"

    # ── Lógica global de agregación (sin cambios) ─────────────────────────

    def test_global_verde_mayoria_cortes_verdes(self, basic_umbrales):
        """Si ≥90% de cortes son VERDE, la propiedad es VERDE."""
        um = {
            ("DENSIDAD", "A"): (2.0, 4.0),
            ("DENSIDAD", "B"): (2.0, 4.0),
            ("DENSIDAD", "C"): (2.0, 4.0),
        }
        errores = {"A": 1.0, "B": 1.0, "C": 1.0}   # todos VERDE
        sem, _, _, _, _, _ = clasificar_propiedad(errores, "DENSIDAD", um, 0.9, 0.3)
        assert sem == "VERDE"

    def test_global_rojo_mayoria_cortes_rojos(self, basic_umbrales):
        """Si >30% de cortes son ROJO, la propiedad es ROJO."""
        um = {
            ("DENSIDAD", "A"): (2.0, 4.0),
            ("DENSIDAD", "B"): (2.0, 4.0),
        }
        errores = {"A": 5.0, "B": 5.0}   # ambos ROJO (≥ 4.0)
        sem, _, _, _, _, _ = clasificar_propiedad(errores, "DENSIDAD", um, 0.9, 0.3)
        assert sem == "ROJO"

    def test_estado_por_corte_correcto(self, basic_umbrales):
        """Verifica el estado individual de cada corte."""
        um = {
            ("DENSIDAD", "150-200"): (2.0, 4.0),
            ("DENSIDAD", "200-250"): (2.0, 4.0),
            ("DENSIDAD", "300-350"): (2.0, 4.0),
        }
        errores = {
            "150-200": 1.0,   # < 2.0 → VERDE
            "200-250": 3.0,   # 2.0 ≤ 3.0 < 4.0 → AMARILLO
            "300-350": 5.0,   # ≥ 4.0 → ROJO
        }
        _, estados, _, _, _, _ = clasificar_propiedad(errores, "DENSIDAD", um, 0.9, 0.3)
        assert estados["150-200"] == "VERDE"
        assert estados["200-250"] == "AMARILLO"
        assert estados["300-350"] == "ROJO"


# ===========================================================================
# 10. Tests _sem_global_por_crudo
# ===========================================================================

class TestSemGlobalPorCrudo:

    def test_todo_verde(self):
        resumen = {
            "DENSIDAD": {"Maya": "VERDE", "Brent": "VERDE"},
            "AZUFRE":   {"Maya": "VERDE", "Brent": "VERDE"},
        }
        out = _sem_global_por_crudo(resumen, 0.9, 0.3)
        assert out["Maya"] == "VERDE"
        assert out["Brent"] == "VERDE"

    def test_mayoria_rojos(self):
        resumen = {
            "DENSIDAD":   {"Maya": "ROJO"},
            "AZUFRE":     {"Maya": "ROJO"},
            "VISCOSIDAD": {"Maya": "VERDE"},
        }
        out = _sem_global_por_crudo(resumen, 0.9, 0.3)
        assert out["Maya"] == "ROJO"   # 2/3 rojos = 0.66 > 0.3

    def test_caso_amarillo(self):
        resumen = {
            "DENSIDAD":   {"Maya": "VERDE"},
            "AZUFRE":     {"Maya": "AMARILLO"},
            "VISCOSIDAD": {"Maya": "VERDE"},
        }
        out = _sem_global_por_crudo(resumen, 0.9, 0.3)
        assert out["Maya"] == "AMARILLO"   # 2/3 verdes = 0.66 < 0.9


# ===========================================================================
# 11. Tests calcular_errores_crudo_df
# ===========================================================================

class TestCalcularErroresCrudoDf:

    def test_basic_output_shape(self, simple_isa, simple_rams, simple_umbrales, alias_prop):
        hoja_resumen = {}
        df_out, cortes, orden = calcular_errores_crudo_df(
            df_isa=simple_isa,
            df_rams=simple_rams,
            umbrales=simple_umbrales,
            alias_prop=alias_prop,
            pct_ok_amarillo=0.9,
            pct_rojo_rojo=0.3,
            hoja_resumen=hoja_resumen,
            crude_name="Maya",
        )
        assert "Semaforo"    in df_out.columns
        assert "Propiedad"   in df_out.columns
        assert "Corte_peor"  in df_out.columns
        assert "Error_peor"  in df_out.columns
        assert "Umbral_peor" in df_out.columns
        assert len(cortes) == 3

    def test_populates_resumen(self, simple_isa, simple_rams, simple_umbrales, alias_prop):
        hoja_resumen = {}
        calcular_errores_crudo_df(
            df_isa=simple_isa, df_rams=simple_rams,
            umbrales=simple_umbrales, alias_prop=alias_prop,
            pct_ok_amarillo=0.9, pct_rojo_rojo=0.3,
            hoja_resumen=hoja_resumen, crude_name="Maya",
        )
        assert len(hoja_resumen) > 0
        for prop_dict in hoja_resumen.values():
            assert "Maya" in prop_dict

    def test_no_propiedad_col_raises(self, alias_prop):
        df = pd.DataFrame({"150-200": [1.0]})
        with pytest.raises(ValueError, match="Propiedad"):
            calcular_errores_crudo_df(
                df_isa=df, df_rams=df,
                umbrales={}, alias_prop=alias_prop,
                pct_ok_amarillo=0.9, pct_rojo_rojo=0.3,
                hoja_resumen={}, crude_name="test",
            )

    def test_errores_son_absolutos(self, alias_prop):
        isa  = pd.DataFrame({"Propiedad": ["Densidad"], "150-200": [850.0]})
        rams = pd.DataFrame({"Propiedad": ["Densidad"], "150-200": [852.0]})
        umbrales = {("DENSIDAD", "150-200"): (5.0, 10.0)}
        df_out, _, _ = calcular_errores_crudo_df(
            df_isa=isa, df_rams=rams,
            umbrales=umbrales, alias_prop=alias_prop,
            pct_ok_amarillo=0.9, pct_rojo_rojo=0.3,
            hoja_resumen={}, crude_name="test",
        )
        assert df_out["150-200"].iloc[0] == pytest.approx(2.0)

    def test_semaforo_verde_cuando_error_menor_repro(self, alias_prop):
        """Error 1.0 < REPRO 2.0 → VERDE."""
        isa  = pd.DataFrame({"Propiedad": ["Densidad"], "150-200": [850.0]})
        rams = pd.DataFrame({"Propiedad": ["Densidad"], "150-200": [851.0]})
        umbrales = {("DENSIDAD", "150-200"): (2.0, 4.0)}
        df_out, _, _ = calcular_errores_crudo_df(
            df_isa=isa, df_rams=rams,
            umbrales=umbrales, alias_prop=alias_prop,
            pct_ok_amarillo=0.9, pct_rojo_rojo=0.3,
            hoja_resumen={}, crude_name="test",
        )
        assert df_out["Semaforo"].iloc[0] == "VERDE"

    def test_semaforo_amarillo_cuando_error_entre_repro_y_admisible(self, alias_prop):
        """Error 3.0: REPRO=2.0 ≤ 3.0 < ADMISIBLE=4.0 → AMARILLO."""
        isa  = pd.DataFrame({"Propiedad": ["Densidad"], "150-200": [850.0]})
        rams = pd.DataFrame({"Propiedad": ["Densidad"], "150-200": [853.0]})
        umbrales = {("DENSIDAD", "150-200"): (2.0, 4.0)}
        df_out, _, _ = calcular_errores_crudo_df(
            df_isa=isa, df_rams=rams,
            umbrales=umbrales, alias_prop=alias_prop,
            pct_ok_amarillo=0.9, pct_rojo_rojo=0.3,
            hoja_resumen={}, crude_name="test",
        )
        assert df_out["Semaforo"].iloc[0] == "AMARILLO"

    def test_semaforo_rojo_cuando_error_mayor_o_igual_admisible(self, alias_prop):
        """Error 4.0 ≥ ADMISIBLE 4.0 → ROJO."""
        isa  = pd.DataFrame({"Propiedad": ["Densidad"], "150-200": [850.0]})
        rams = pd.DataFrame({"Propiedad": ["Densidad"], "150-200": [854.0]})
        umbrales = {("DENSIDAD", "150-200"): (2.0, 4.0)}
        df_out, _, _ = calcular_errores_crudo_df(
            df_isa=isa, df_rams=rams,
            umbrales=umbrales, alias_prop=alias_prop,
            pct_ok_amarillo=0.9, pct_rojo_rojo=0.3,
            hoja_resumen={}, crude_name="test",
        )
        assert df_out["Semaforo"].iloc[0] == "ROJO"


# ===========================================================================
# 12. Tests validate_params
# ===========================================================================

class TestValidateParams:

    def test_valid(self):
        validate_params(0.1, 0.6, 0.9, 0.3)  # sin excepción

    def test_negative_tol(self):
        with pytest.raises(ValueError, match="negativo"):
            validate_params(-0.1, 0.6, 0.9, 0.3)

    def test_pct_out_of_range(self):
        with pytest.raises(ValueError, match="0,1"):
            validate_params(0.1, 0.6, 1.5, 0.3)


# ===========================================================================
# 13. Tests build_summary_df
# ===========================================================================

class TestBuildSummaryDf:

    def test_global_row_first(self):
        resumen = {
            "DENSIDAD": {"Maya": "VERDE"},
            "AZUFRE":   {"Maya": "ROJO"},
        }
        df = _build_summary_df(resumen, ["DENSIDAD", "AZUFRE"], 0.9, 0.3)
        assert df.iloc[0]["Propiedad"] == "GLOBAL"

    def test_all_props_present(self):
        resumen = {
            "DENSIDAD": {"Maya": "VERDE"},
            "AZUFRE":   {"Maya": "ROJO"},
        }
        df = _build_summary_df(resumen, ["DENSIDAD", "AZUFRE"], 0.9, 0.3)
        propiedades = df["Propiedad"].tolist()
        assert "DENSIDAD" in propiedades
        assert "AZUFRE" in propiedades

    def test_crudo_columns(self):
        resumen = {"DENSIDAD": {"Maya": "VERDE", "Brent": "ROJO"}}
        df = _build_summary_df(resumen, ["DENSIDAD"], 0.9, 0.3)
        assert "Maya" in df.columns
        assert "Brent" in df.columns


# ===========================================================================
# 14. Tests build_excel
# ===========================================================================

class TestBuildExcelMVP:

    def _make_result(self, alias_prop, simple_isa, simple_rams, simple_umbrales):
        hoja_resumen = {}
        df_out, cortes, orden = calcular_errores_crudo_df(
            df_isa=simple_isa, df_rams=simple_rams,
            umbrales=simple_umbrales, alias_prop=alias_prop,
            pct_ok_amarillo=0.9, pct_rojo_rojo=0.3,
            hoja_resumen=hoja_resumen, crude_name="Maya",
        )
        summary = _build_summary_df(hoja_resumen, orden, 0.9, 0.3)
        return ValidationResult(
            paired_names=["Maya"],
            crudo_dataframes={"Maya": df_out},
            cortes_visibles={"Maya": cortes},
            resumen_raw=hoja_resumen,
            orden_propiedades=orden,
            summary=summary,
        )

    def test_returns_bytes(self, alias_prop, simple_isa, simple_rams, simple_umbrales):
        result = self._make_result(alias_prop, simple_isa, simple_rams, simple_umbrales)
        excel  = build_excel(result)
        assert isinstance(excel, bytes) and len(excel) > 0

    def test_has_resumen_sheet(self, alias_prop, simple_isa, simple_rams, simple_umbrales):
        import openpyxl
        result = self._make_result(alias_prop, simple_isa, simple_rams, simple_umbrales)
        wb = openpyxl.load_workbook(io.BytesIO(build_excel(result)))
        assert "Resumen" in wb.sheetnames

    def test_has_crudo_sheet(self, alias_prop, simple_isa, simple_rams, simple_umbrales):
        import openpyxl
        result = self._make_result(alias_prop, simple_isa, simple_rams, simple_umbrales)
        wb = openpyxl.load_workbook(io.BytesIO(build_excel(result)))
        assert "Maya" in wb.sheetnames

    def test_resumen_has_global_row(self, alias_prop, simple_isa, simple_rams, simple_umbrales):
        import openpyxl
        result = self._make_result(alias_prop, simple_isa, simple_rams, simple_umbrales)
        wb = openpyxl.load_workbook(io.BytesIO(build_excel(result)))
        ws = wb["Resumen"]
        assert ws.cell(2, 1).value == "GLOBAL"

    def test_crudo_sheet_has_semaforo_col(self, alias_prop, simple_isa, simple_rams, simple_umbrales):
        import openpyxl
        result = self._make_result(alias_prop, simple_isa, simple_rams, simple_umbrales)
        wb = openpyxl.load_workbook(io.BytesIO(build_excel(result)))
        ws = wb["Maya"]
        headers = [ws.cell(1, j).value for j in range(1, ws.max_column + 1)]
        assert "Semaforo" in headers


# ===========================================================================
# 15. Tests run_validation (pipeline completo)
# ===========================================================================

class TestRunValidation:

    def test_full_pipeline(self, isa_bytes, rams_bytes, matriz_bytes):
        result = run_validation(
            isa_files={"ISA_Maya.xlsx": io.BytesIO(isa_bytes)},
            rams_files={"RAMS_Maya.xlsx": io.BytesIO(rams_bytes)},
            matriz_file=io.BytesIO(matriz_bytes),
            matriz_filename="Errores_Cortes.xlsx",
        )
        assert result.has_results
        assert "Maya" in result.paired_names

    def test_pipeline_no_pairs(self, isa_bytes, rams_bytes, matriz_bytes):
        result = run_validation(
            isa_files={"ISA_Maya.xlsx": io.BytesIO(isa_bytes)},
            rams_files={"RAMS_Brent.xlsx": io.BytesIO(rams_bytes)},
            matriz_file=io.BytesIO(matriz_bytes),
            matriz_filename="Errores_Cortes.xlsx",
        )
        assert not result.has_results

    def test_pipeline_summary_has_global(self, isa_bytes, rams_bytes, matriz_bytes):
        result = run_validation(
            isa_files={"ISA_Maya.xlsx": io.BytesIO(isa_bytes)},
            rams_files={"RAMS_Maya.xlsx": io.BytesIO(rams_bytes)},
            matriz_file=io.BytesIO(matriz_bytes),
            matriz_filename="Errores_Cortes.xlsx",
        )
        assert "GLOBAL" in result.summary["Propiedad"].values

    def test_pipeline_crudo_dataframe_cols(self, isa_bytes, rams_bytes, matriz_bytes):
        result = run_validation(
            isa_files={"ISA_Maya.xlsx": io.BytesIO(isa_bytes)},
            rams_files={"RAMS_Maya.xlsx": io.BytesIO(rams_bytes)},
            matriz_file=io.BytesIO(matriz_bytes),
            matriz_filename="Errores_Cortes.xlsx",
        )
        df = result.crudo_dataframes["Maya"]
        for col in ["Propiedad", "Semaforo", "Corte_peor", "Error_peor", "Umbral_peor"]:
            assert col in df.columns, f"Falta columna: {col}"

    def test_invalid_tol_raises(self, isa_bytes, rams_bytes, matriz_bytes):
        with pytest.raises(ValueError):
            run_validation(
                isa_files={"ISA_Maya.xlsx": io.BytesIO(isa_bytes)},
                rams_files={"RAMS_Maya.xlsx": io.BytesIO(rams_bytes)},
                matriz_file=io.BytesIO(matriz_bytes),
                matriz_filename="Errores_Cortes.xlsx",
                tol=-0.1,
            )

    def test_clasificacion_verde_con_error_menor_repro(self, isa_bytes, matriz_bytes):
        """Pipeline end-to-end: error pequeño → VERDE en el resumen."""
        # ISA y RAMS casi idénticos → error ≈ 0 < REPRO → todos verdes
        isa_df  = pd.DataFrame({
            "Propiedad": ["Densidad"], "150-200": [850.0], "200-250": [860.0], "300-350": [870.0]
        })
        rams_df = pd.DataFrame({
            "Propiedad": ["Densidad"], "150-200": [850.1], "200-250": [860.1], "300-350": [870.1]
        })
        buf_isa  = io.BytesIO(); isa_df.to_excel(buf_isa,   index=False); buf_isa.seek(0)
        buf_rams = io.BytesIO(); rams_df.to_excel(buf_rams, index=False); buf_rams.seek(0)

        result = run_validation(
            isa_files={"ISA_T.xlsx": buf_isa},
            rams_files={"RAMS_T.xlsx": buf_rams},
            matriz_file=io.BytesIO(matriz_bytes),
            matriz_filename="Errores_Cortes.xlsx",
        )
        sem_densidad = result.resumen_raw.get("DENSIDAD", {}).get("T")
        assert sem_densidad == "VERDE"


# ===========================================================================
# 16. Tests de compatibilidad
# ===========================================================================

class TestBackwardsCompat:

    def test_canonize_name_alias(self):
        assert canonize_name("ISA_Maya.xlsx") == _nombre_base_crudo("ISA_Maya.xlsx")

    def test_validate_thresholds_valid(self):
        config = ThresholdConfig(default_green=1.0, default_yellow=3.0)
        validate_thresholds(config)

    def test_validate_thresholds_invalid(self):
        config = ThresholdConfig(default_green=5.0, default_yellow=3.0)
        with pytest.raises(ValueError, match="globales"):
            validate_thresholds(config)

    def test_validate_thresholds_per_prop(self):
        config = ThresholdConfig(
            default_green=1.0,
            default_yellow=3.0,
            thresholds={"densidad": (5.0, 2.0)},
        )
        with pytest.raises(ValueError, match="densidad"):
            validate_thresholds(config)

    def test_constants_present(self):
        assert "VERDE"    in SEMAFORO_COLORS
        assert "AMARILLO" in SEMAFORO_COLORS
        assert "ROJO"     in SEMAFORO_COLORS
        assert ".xlsx"    in SUPPORTED_EXTENSIONS
        assert ".csv"     in SUPPORTED_EXTENSIONS

    def test_default_constants(self):
        assert DEFAULT_TOL             == 0.10
        assert DEFAULT_TOL_PESADOS     == 0.60
        assert DEFAULT_PCT_OK_AMARILLO == 0.90
        assert DEFAULT_PCT_ROJO_ROJO   == 0.30
