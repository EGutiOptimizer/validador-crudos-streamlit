"""
core/validator_core.py
======================
Lógica de negocio pura para validación de crudos RAMS vs ISA.

LÓGICA DE SEMÁFORO POR CORTE:
  - error < REPRO                → 🟢 VERDE
  - REPRO ≤ error < ADMISIBLE   → 🟡 AMARILLO
  - error ≥ ADMISIBLE            → 🔴 ROJO
  - Sin umbral definido          → ⚪ N/A

LÓGICA GLOBAL POR PROPIEDAD (sin cambios):
  - ≥ pct_ok_amarillo de cortes VERDE → VERDE global
  - > pct_rojo_rojo  de cortes ROJO   → ROJO global
  - resto                             → AMARILLO global

Sin imports de Streamlit — 100 % testeable en aislamiento.
"""
from __future__ import annotations

import io
import logging
import re
import unicodedata
from typing import IO, Any, Dict, List, Optional, Tuple

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter

from core.models import ValidationResult, ThresholdConfig

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constantes
# ---------------------------------------------------------------------------

SEMAFORO_COLORS = {
    "VERDE":    "C6EFCE",
    "AMARILLO": "FFEB9C",
    "ROJO":     "FFC7CE",
    "NA":       "E7E6E6",
}

SEMAFORO_LABELS = {
    "VERDE":    "🟢 OK",
    "AMARILLO": "🟡 Revisar",
    "ROJO":     "🔴 Fuera de rango",
    "NA":       "⚪ Sin umbral",
}

SUPPORTED_EXTENSIONS = {".xlsx", ".xls", ".csv"}

# Mantenidas por compatibilidad con código externo / tests existentes
# (ya no se usan internamente en la clasificación por corte)
DEFAULT_TOL             = 0.10
DEFAULT_TOL_PESADOS     = 0.60
DEFAULT_PCT_OK_AMARILLO = 0.90
DEFAULT_PCT_ROJO_ROJO   = 0.30

# Tipo alias para el diccionario de umbrales
# { (PROP_CANON, CORTE_CANON): (repro, admisible) }
# Cualquiera de los dos puede ser None si no figura en la matriz
UmbralesDict = Dict[Tuple[str, str], Tuple[Optional[float], Optional[float]]]


# ---------------------------------------------------------------------------
# 1. Normalización de texto
# ---------------------------------------------------------------------------

def strip_accents(text: str) -> str:
    if text is None:
        return ""
    text = str(text)
    return "".join(
        c for c in unicodedata.normalize("NFD", text)
        if unicodedata.category(c) != "Mn"
    )


def _canon_prop_norm(s: str) -> str:
    if s is None:
        return ""
    t = strip_accents(str(s)).upper().strip()
    for ch in [".", "º", "°"]:
        t = t.replace(ch, "")
    for ch in ["(", ")", "%", "/"]:
        t = t.replace(ch, " ")
    t = re.sub(r"\s*,\s*", " ", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t


def canon_prop(s: str, alias: Optional[Dict[str, str]] = None) -> str:
    t = _canon_prop_norm(s)
    if not re.search(r"[A-Z0-9]", t):
        return ""
    if alias:
        return alias.get(t, t)
    return t


def canon_corte(s: str) -> str:
    if s is None:
        return ""
    t = str(s)
    t = t.replace("\u00A0", " ")
    t = "".join(" " if unicodedata.category(c) == "Zs" else c for c in t)
    for dash in ["\u2010", "\u2011", "\u2012", "\u2013", "\u2014", "\u2212"]:
        t = t.replace(dash, "-")
    t = t.replace("º", "").replace("°", "")
    t = re.sub(r"\s*-\s*", "-", t)
    t = re.sub(r"\s+", "", t).upper()
    return t


# ---------------------------------------------------------------------------
# 2. Construcción de umbrales
#    Ahora almacena REPRO y ADMISIBLE por separado en una tupla (repro, admis).
# ---------------------------------------------------------------------------

def detectar_columna_tipo(df: pd.DataFrame) -> Optional[str]:
    for col in df.columns:
        name = str(col).strip().lower()
        if name in {"tipo", "columna1", "categoria", "categoría"}:
            return col
    sample_cols = [c for c in df.columns if str(c).strip().lower() not in {"propiedad", "unidad"}]
    for col in sample_cols:
        try:
            serie = df[col].astype(str).str.upper().str.strip()
            if serie.head(50).str.contains(r"REPRO|ADMISIBLE|REPET").any():
                return col
        except Exception:
            pass
    return None


def normalizar_tipo(raw) -> str:
    if pd.isna(raw):
        return ""
    t = str(raw)
    t = t.replace("*", "")
    t = t.replace("\u00A0", " ")
    t = strip_accents(t).upper().strip()
    return t


def construir_umbrales(
    df: pd.DataFrame,
    alias_prop: Dict[str, str],
) -> UmbralesDict:
    """
    Construye el diccionario de umbrales leyendo REPRO y ADMISIBLE por separado.

    Devuelve:
        { (PROP_CANON, CORTE_CANON): (repro_or_None, admisible_or_None) }

    Reglas:
      - Filas con Tipo que contenga REPRO o REPET → alimentan repro_dict
      - Filas con Tipo que contenga ADMISIBLE     → alimentan admis_dict
      - Si hay varios valores del mismo tipo para la misma clave, se conserva el mayor.
      - Las claves que aparezcan en cualquiera de los dos dicts forman el resultado.
    """
    col_prop = next((c for c in df.columns if str(c).strip().lower() == "propiedad"), None)
    if col_prop is None:
        raise ValueError("La matriz de umbrales no tiene columna 'Propiedad'.")
    col_tipo = detectar_columna_tipo(df)
    if col_tipo is None:
        raise ValueError("No se localiza columna 'Tipo' en la matriz de umbrales.")

    cortes_cols: List[Tuple[str, str]] = []
    for c in df.columns:
        if c in {col_prop, col_tipo}:
            continue
        cc = canon_corte(str(c))
        if cc in {"", "UNIDAD", "CRUDO"}:
            continue
        cortes_cols.append((str(c), cc))

    repro_dict: Dict[Tuple[str, str], float] = {}
    admis_dict: Dict[Tuple[str, str], float] = {}
    prop_actual = ""

    for _, row in df.iterrows():
        prop_raw = row.get(col_prop)
        prop_c = canon_prop(prop_raw, alias_prop)
        if prop_c:
            prop_actual = prop_c
        if not prop_actual:
            continue

        tipo = normalizar_tipo(row.get(col_tipo))
        is_repro = "REPRO" in tipo or "REPET" in tipo
        is_admis = "ADMISIBLE" in tipo

        if not (is_repro or is_admis):
            continue

        for col_o, cc in cortes_cols:
            val = row.get(col_o)
            if pd.isna(val):
                continue
            vs = str(val).strip()
            if vs == "":
                continue
            try:
                v = float(vs.replace(",", "."))
            except Exception:
                continue

            key = (prop_actual, cc)
            if is_repro:
                if key not in repro_dict or v > repro_dict[key]:
                    repro_dict[key] = v
            if is_admis:
                if key not in admis_dict or v > admis_dict[key]:
                    admis_dict[key] = v

    all_keys = set(repro_dict.keys()) | set(admis_dict.keys())
    return {k: (repro_dict.get(k), admis_dict.get(k)) for k in all_keys}


# ---------------------------------------------------------------------------
# 3. Semántica de alias de propiedades  (sin cambios)
# ---------------------------------------------------------------------------

def crear_semantica_alias() -> Dict[str, str]:
    raw: Dict[str, str] = {
        # ── Peso ─────────────────────────────────────────────────────────────
        "PESO":                             "PESO",
        "PESO ACUMULADO":                   "PESO ACUMULADO",
        "RENDIMIENTO":                      "PESO",
        "RENDIMIENTO ACUMULADO":            "PESO ACUMULADO",
        "% DESTILADO":                      "PESO",
        "% VOL":                            "PESO",
        "% EN PESO":                        "PESO",
        # ── Densidad ─────────────────────────────────────────────────────────
        "DENSIDAD":                         "DENSIDAD",
        "DENSIDAD A 15C":                   "DENSIDAD",
        "DENSIDAD A 15":                    "DENSIDAD",
        "DENSIDAD 15C":                     "DENSIDAD",
        "DENSIDAD RELATIVA":                "DENSIDAD",
        "DENSIDAD RELATIVA 15/4":           "DENSIDAD",
        "DENSIDAD A 15/4":                  "DENSIDAD",
        "D15":                              "DENSIDAD",
        "GRAVEDAD ESPECIFICA":              "DENSIDAD",
        # ── Viscosidad ───────────────────────────────────────────────────────
        "VISCOSIDAD 50":                    "VISCOSIDAD 50",
        "VISCOSIDAD 50C":                   "VISCOSIDAD 50",
        "VISCOSIDAD A 50C":                 "VISCOSIDAD 50",
        "VISCOSIDAD CINEMATICA 50":         "VISCOSIDAD 50",
        "VISCOSIDAD CINEMATICA 50C":        "VISCOSIDAD 50",
        "VISCOSIDAD DINAMICA 50":           "VISCOSIDAD 50",
        "VISCOSIDAD DINAMICA 50C":          "VISCOSIDAD 50",
        "VIS 50":                           "VISCOSIDAD 50",
        "VISCOSIDAD 100":                   "VISCOSIDAD 100",
        "VISCOSIDAD 100C":                  "VISCOSIDAD 100",
        "VISCOSIDAD A 100C":                "VISCOSIDAD 100",
        "VISCOSIDAD CINEMATICA 100":        "VISCOSIDAD 100",
        "VISCOSIDAD CINEMATICA 100C":       "VISCOSIDAD 100",
        "VIS 100":                          "VISCOSIDAD 100",
        # ── Azufre ───────────────────────────────────────────────────────────
        "AZUFRE":                           "AZUFRE",
        "AZUFRE TOTAL":                     "AZUFRE",
        "AZUFRE MERCAPTANO":                "AZUFRE MERCAPTANO",
        "S MERCAPTANO":                     "AZUFRE MERCAPTANO",
        # ── Octanaje ─────────────────────────────────────────────────────────
        "RON":                              "RON",
        "NOR":                              "RON",
        "NOR CLARO":                        "RON",
        "N O R CLARO":                      "RON",
        "NUMERO DE OCTANO INVESTIGACION":   "RON",
        "MON":                              "MON",
        "NOM":                              "MON",
        "NOM CLARO":                        "MON",
        "N O M CLARO":                      "MON",
        "NUMERO DE OCTANO MOTOR":           "MON",
        # ── Índice de neutralización ─────────────────────────────────────────
        "N DE NEUTRALIZACION":              "N DE NEUTRALIZACION",
        "NUMERO DE NEUTRALIZACION":         "N DE NEUTRALIZACION",
        "NO DE NEUTRALIZACION":             "N DE NEUTRALIZACION",
        "INDICE DE ACIDEZ":                 "N DE NEUTRALIZACION",
        # ── Índice de refracción ─────────────────────────────────────────────
        "INDICE DE REFRACCION 70C":         "INDICE DE REFRACCION 70C",
        "INDICE DE REFRACCION":             "INDICE DE REFRACCION 70C",
        "IR 70C":                           "INDICE DE REFRACCION 70C",
        # ── Puntos físicos ───────────────────────────────────────────────────
        "PUNTO DE VERTIDO":                 "PUNTO DE VERTIDO",
        "PUNTO DE NIEBLA":                  "PUNTO DE NIEBLA",
        "PUNTO DE CRISTALIZACION":          "PUNTO DE CRISTALIZACION",
        "PUNTO DE ANILINA":                 "PUNTO DE ANILINA",
        "PUNTO DE INFLAMACION":             "PUNTO DE INFLAMACION",
        "PUNTO INICIAL DE EBULLICION":      "PUNTO INICIAL DE EBULLICION",
        "PIE":                              "PUNTO INICIAL DE EBULLICION",
        "IBP":                              "PUNTO INICIAL DE EBULLICION",
        "PUNTO FINAL DE EBULLICION":        "PUNTO FINAL DE EBULLICION",
        "PFE":                              "PUNTO FINAL DE EBULLICION",
        "FBP":                              "PUNTO FINAL DE EBULLICION",
        # ── PIONA ────────────────────────────────────────────────────────────
        "PIONA (%VOL), N-PARAFINAS":        "PIONA N-PARAFINAS",
        "PIONA (%VOL), I-PARAFINAS":        "PIONA I-PARAFINAS",
        "PIONA (%VOL), NAFTENOS":           "PIONA NAFTENOS",
        "PIONA (%VOL), POLINAFTENOS":       "PIONA POLINAFTENOS",
        "PIONA (%VOL), AROMATICOS":         "PIONA AROMATICOS",
        "PIONA (%VOL), OLEFINAS":           "PIONA OLEFINAS",
        "PIONA (%VOL), SUPERIORES A 200C":  "PIONA SUPERIORES A 200C",
        "PIONA (%VOL) N-PARAFINAS":         "PIONA N-PARAFINAS",
        "PIONA (%VOL) I-PARAFINAS":         "PIONA I-PARAFINAS",
        "PIONA (%VOL) NAFTENOS":            "PIONA NAFTENOS",
        "PIONA (%VOL) POLINAFTENOS":        "PIONA POLINAFTENOS",
        "PIONA (%VOL) AROMATICOS":          "PIONA AROMATICOS",
        "PIONA (%VOL) OLEFINAS":            "PIONA OLEFINAS",
        "PIONA (%VOL) SUPERIORES A 200C":   "PIONA SUPERIORES A 200C",
        "PIONA N-PARAFINAS":                "PIONA N-PARAFINAS",
        "PIONA I-PARAFINAS":                "PIONA I-PARAFINAS",
        "PIONA NAFTENOS":                   "PIONA NAFTENOS",
        "PIONA POLINAFTENOS":               "PIONA POLINAFTENOS",
        "PIONA AROMATICOS":                 "PIONA AROMATICOS",
        "PIONA OLEFINAS":                   "PIONA OLEFINAS",
        "PIONA SUPERIORES A 200C":          "PIONA SUPERIORES A 200C",
        "PIONA, N-PARAFINAS":               "PIONA N-PARAFINAS",
        "PIONA, I-PARAFINAS":               "PIONA I-PARAFINAS",
        "PIONA, NAFTENOS":                  "PIONA NAFTENOS",
        "PIONA, POLINAFTENOS":              "PIONA POLINAFTENOS",
        "PIONA, AROMATICOS":                "PIONA AROMATICOS",
        "PIONA, OLEFINAS":                  "PIONA OLEFINAS",
        "PIONA, SUPERIORES A 200C":         "PIONA SUPERIORES A 200C",
        "N-PARAFINAS":                      "PIONA N-PARAFINAS",
        "PARAFINAS NORMALES":               "PIONA N-PARAFINAS",
        "N PARAFINAS":                      "PIONA N-PARAFINAS",
        "I-PARAFINAS":                      "PIONA I-PARAFINAS",
        "ISOPARAFINAS":                     "PIONA I-PARAFINAS",
        "I PARAFINAS":                      "PIONA I-PARAFINAS",
        "NAFTENOS":                         "PIONA NAFTENOS",
        "NAFTENICOS":                       "PIONA NAFTENOS",
        "POLINAFTENOS":                     "PIONA POLINAFTENOS",
        "AROMATICOS":                       "PIONA AROMATICOS",
        "AROMATICS":                        "PIONA AROMATICOS",
        "OLEFINAS":                         "PIONA OLEFINAS",
        "SUPERIORES A 200C":                "PIONA SUPERIORES A 200C",
        # ── Nitrógeno ────────────────────────────────────────────────────────
        "NITROGENO":                        "NITROGENO",
        "NITROGENO TOTAL":                  "NITROGENO",
        "NITROGENO BASICO":                 "NITROGENO BASICO",
        # ── Residuo de carbono ───────────────────────────────────────────────
        "RESIDUO DE CARBON":                "RESIDUO DE CARBON",
        "CARBONO CONRADSON":                "RESIDUO DE CARBON",
        "CONRADSON":                        "RESIDUO DE CARBON",
        "CCR":                              "RESIDUO DE CARBON",
        "MCRT":                             "RESIDUO DE CARBON",
        # ── Asfaltenos y aromáticos ──────────────────────────────────────────
        "ASFALTENOS":                       "ASFALTENOS",
        "MONOAROMATICOS":                   "MONOAROMATICOS",
        "DIAROMATICOS":                     "DIAROMATICOS",
        "TRIAROMATICOS Y SUPERIORES":       "TRIAROMATICOS",
        "TRIAROMATICOS":                    "TRIAROMATICOS",
        # ── Gases ligeros ────────────────────────────────────────────────────
        "CONTENIDO EN C2":                  "CONTENIDO EN C2",
        "C2":                               "CONTENIDO EN C2",
        "CONTENIDO EN C3":                  "CONTENIDO EN C3",
        "C3":                               "CONTENIDO EN C3",
        "CONTENIDO EN IC4":                 "CONTENIDO EN IC4",
        "IC4":                              "CONTENIDO EN IC4",
        "CONTENIDO EN NC4":                 "CONTENIDO EN NC4",
        "NC4":                              "CONTENIDO EN NC4",
        # ── Metales ──────────────────────────────────────────────────────────
        "NIQUEL":                           "NIQUEL",
        "NI":                               "NIQUEL",
        "VANADIO":                          "VANADIO",
        "V":                                "VANADIO",
        "SILICIO":                          "SILICIO",
        "SI":                               "SILICIO",
    }

    out: Dict[str, str] = {}
    for k, v in raw.items():
        nk = _canon_prop_norm(k)
        nv = _canon_prop_norm(v)
        if nk:
            out[nk] = nv
    return out


# ---------------------------------------------------------------------------
# 4. Helpers de umbral — REPRO/ADMISIBLE
# ---------------------------------------------------------------------------

def es_corte_pesado(corte: str) -> bool:
    """Mantenida por compatibilidad; ya no se usa internamente en clasificación."""
    s = corte.upper().strip()
    if any(tag in s for tag in ["C6", "C7", "C8", "C9", "C10"]):
        return True
    if s.endswith("+"):
        try:
            return float(s[:-1]) >= 299
        except Exception:
            return False
    if "-" in s:
        try:
            return float(s.split("-")[0]) >= 299
        except Exception:
            return False
    return False


def _prop_base_para_umbral(prop_canon: str) -> str:
    if prop_canon == "PESO ACUMULADO":
        return "PESO"
    return prop_canon


def _buscar_umbrales(
    umbrales: UmbralesDict,
    prop_canon: str,
    corte_key: str,
) -> Tuple[Optional[float], Optional[float]]:
    """
    Devuelve (repro, admisible) para la combinación propiedad/corte dada.
    Aplica fallback de PESO ACUMULADO → PESO si no hay umbral directo.
    """
    result = umbrales.get((prop_canon, corte_key))
    if result is not None:
        return result
    base_prop = _prop_base_para_umbral(prop_canon)
    if base_prop != prop_canon:
        result2 = umbrales.get((base_prop, corte_key))
        if result2 is not None:
            return result2
    return (None, None)


# Alias de compatibilidad: devuelve el máximo de (repro, admis) como antes
def _buscar_umbral(
    umbrales: UmbralesDict,
    prop_canon: str,
    corte_key: str,
) -> Optional[float]:
    repro, admis = _buscar_umbrales(umbrales, prop_canon, corte_key)
    vals = [v for v in (repro, admis) if v is not None]
    return max(vals) if vals else None


# ---------------------------------------------------------------------------
# 5. Clasificación por corte y semáforo global de propiedad
#
#  NUEVA LÓGICA POR CORTE:
#    error < REPRO                → VERDE
#    REPRO ≤ error < ADMISIBLE   → AMARILLO
#    error ≥ ADMISIBLE            → ROJO
#    Sin REPRO pero con ADMISIBLE: error < ADMISIBLE → VERDE, ≥ → ROJO
#    Sin ninguno                  → "(sin umbral)"
#
#  LÓGICA GLOBAL (sin cambios):
#    % verdes ≥ pct_ok_amarillo  → VERDE
#    % rojos  > pct_rojo_rojo    → ROJO
#    resto                       → AMARILLO
# ---------------------------------------------------------------------------

def clasificar_propiedad(
    errores_fila: Dict[str, float],
    prop_canon: str,
    umbrales: UmbralesDict,
    pct_ok_amarillo: float,
    pct_rojo_rojo: float,
    # Los siguientes parámetros se mantienen en la firma por compatibilidad
    # con código externo pero ya no se usan en la lógica de clasificación.
    tol: float = DEFAULT_TOL,
    tol_pesados: float = DEFAULT_TOL_PESADOS,
) -> Tuple[str, Dict[str, str], Optional[str], Optional[float], float, Optional[float]]:
    """
    Clasifica una propiedad completa (todos sus cortes).

    Returns:
        (semáforo_global, estados_por_corte, corte_peor, error_peor,
         ratio_peor, umbral_repro_peor)
    """
    estados: Dict[str, str] = {}
    total_con_valor = 0
    total_valid = 0
    n_verde = n_amarillo = n_rojo = 0

    corte_peor: Optional[str] = None
    error_peor: Optional[float] = None
    umbral_peor: Optional[float] = None   # REPRO del corte más crítico
    ratio_peor = -1.0

    base_prop = _prop_base_para_umbral(prop_canon)
    tiene_umbral_prop = any(
        k[0] in (prop_canon, base_prop)
        for k in umbrales.keys()
    )

    for corte, valor in errores_fila.items():
        if valor is None or (isinstance(valor, float) and pd.isna(valor)):
            estados[corte] = "(no numérico)"
            continue

        try:
            v = float(str(valor).replace(",", "."))
        except Exception:
            estados[corte] = "(no numérico)"
            continue

        total_con_valor += 1

        cc = canon_corte(corte)
        repro, admis = _buscar_umbrales(umbrales, prop_canon, cc)
        if repro is None and admis is None:
            # Intentar con el nombre de corte sin canonizar (fallback)
            repro, admis = _buscar_umbrales(umbrales, prop_canon, corte)

        if repro is None and admis is None:
            estados[corte] = "(sin umbral)"
            continue

        total_valid += 1

        # Referencia para el "peor corte": ratio respecto a REPRO (o ADMISIBLE si no hay REPRO)
        ref_thr = repro if repro is not None else admis
        try:
            ratio = v / ref_thr
        except ZeroDivisionError:
            ratio = float("inf")

        if ratio > ratio_peor:
            ratio_peor = ratio
            corte_peor = corte
            error_peor = v
            umbral_peor = repro  # Guardamos el REPRO como referencia

        # ── Clasificación REPRO / ADMISIBLE ──────────────────────────────────
        if repro is not None and v < repro:
            estados[corte] = "VERDE"
            n_verde += 1
        elif admis is not None and v < admis:
            # v ≥ repro (o repro es None) y v < admis
            estados[corte] = "AMARILLO"
            n_amarillo += 1
        else:
            # v ≥ admis, o v ≥ repro sin admis definida
            estados[corte] = "ROJO"
            n_rojo += 1

    # ── Semáforo global de la propiedad ──────────────────────────────────────
    if total_con_valor == 0:
        return "", estados, corte_peor, error_peor, ratio_peor, umbral_peor

    if not tiene_umbral_prop or total_valid == 0:
        return "NA", estados, corte_peor, error_peor, ratio_peor, umbral_peor

    verde_pct = n_verde / total_valid
    rojo_pct  = n_rojo  / total_valid

    if rojo_pct > pct_rojo_rojo:
        return "ROJO", estados, corte_peor, error_peor, ratio_peor, umbral_peor
    if verde_pct >= pct_ok_amarillo:
        return "VERDE", estados, corte_peor, error_peor, ratio_peor, umbral_peor

    return "AMARILLO", estados, corte_peor, error_peor, ratio_peor, umbral_peor


# ---------------------------------------------------------------------------
# 6. Semáforo global por crudo  (sin cambios)
# ---------------------------------------------------------------------------

def _sem_global_por_crudo(
    resumen: Dict[str, Dict[str, str]],
    pct_ok_amarillo: float,
    pct_rojo_rojo: float,
) -> Dict[str, str]:
    crudos = sorted({c for m in resumen.values() for c in m.keys()})
    out: Dict[str, str] = {c: "" for c in crudos}
    for cr in crudos:
        vals = [m.get(cr, "") for m in resumen.values()]
        vals = [v for v in vals if v in {"VERDE", "AMARILLO", "ROJO"}]
        if not vals:
            out[cr] = ""
            continue
        total = len(vals)
        verde = sum(1 for v in vals if v == "VERDE")
        rojo  = sum(1 for v in vals if v == "ROJO")
        if rojo / total > pct_rojo_rojo:
            out[cr] = "ROJO"
        elif verde / total >= pct_ok_amarillo:
            out[cr] = "VERDE"
        else:
            out[cr] = "AMARILLO"
    return out


# ---------------------------------------------------------------------------
# 7. Lectura de archivos en memoria
# ---------------------------------------------------------------------------

def _get_extension(filename: str) -> str:
    match = re.search(r"(\.[^.]+)$", filename.lower())
    return match.group(1) if match else ""


def _read_csv_mem(data: io.BytesIO, filename: str) -> pd.DataFrame:
    for sep in [";", ",", "\t", "|"]:
        data.seek(0)
        try:
            df = pd.read_csv(data, sep=sep, decimal=",", thousands=".", encoding="utf-8-sig")
            if df.shape[1] > 1:
                return df
        except Exception:
            continue
    data.seek(0)
    try:
        return pd.read_csv(data, sep=None, engine="python", decimal=",", encoding="utf-8-sig")
    except Exception as e:
        raise ValueError(f"No se pudo parsear CSV '{filename}': {e}") from e


def read_file(file_obj: IO[bytes], filename: str) -> pd.DataFrame:
    ext = _get_extension(filename)
    if ext not in SUPPORTED_EXTENSIONS:
        raise ValueError(f"Formato no soportado: '{ext}'. Use: {', '.join(sorted(SUPPORTED_EXTENSIONS))}")

    data = io.BytesIO(file_obj.read() if hasattr(file_obj, "read") else file_obj)

    if ext == ".csv":
        return _read_csv_mem(data, filename)
    elif ext == ".xlsx":
        try:
            return pd.read_excel(data, engine="openpyxl")
        except Exception as e:
            raise ValueError(f"Error leyendo '{filename}': {e}") from e
    elif ext == ".xls":
        try:
            return pd.read_excel(data, engine="xlrd")
        except Exception as e:
            raise ValueError(f"Error leyendo '{filename}': {e}") from e
    else:
        raise ValueError(f"Extensión inesperada: {ext}")


def read_file_with_sheet(
    file_obj: IO[bytes],
    filename: str,
    sheet_hint: Optional[str] = None,
) -> pd.DataFrame:
    ext = _get_extension(filename)
    data = io.BytesIO(file_obj.read() if hasattr(file_obj, "read") else file_obj)

    if ext == ".csv":
        return _read_csv_mem(data, filename)
    elif ext in (".xlsx", ".xls"):
        engine = "openpyxl" if ext == ".xlsx" else "xlrd"
        sheet = sheet_hint or 0
        try:
            return pd.read_excel(data, sheet_name=sheet, engine=engine)
        except Exception:
            data.seek(0)
            alt_engine = "xlrd" if engine == "openpyxl" else "openpyxl"
            try:
                return pd.read_excel(data, sheet_name=sheet, engine=alt_engine)
            except Exception as e:
                raise ValueError(f"No se pudo leer '{filename}': {e}") from e
    else:
        raise ValueError(f"Formato no soportado: '{ext}'")


# ---------------------------------------------------------------------------
# 8. Detectar cortes y helpers de índices
# ---------------------------------------------------------------------------

def detectar_cortes_en_df(df: pd.DataFrame) -> List[Tuple[str, str]]:
    cortes = []
    META = {"propiedad", "unidad", "validacion", "validación", "validacion auto", "validación auto"}
    for c in df.columns:
        if str(c).strip().lower() in META:
            continue
        cc = canon_corte(str(c))
        if cc:
            cortes.append((str(c), cc))
    return cortes


def _indice_prop(df: pd.DataFrame, alias_prop: Dict[str, str]) -> Dict[str, int]:
    idx: Dict[str, int] = {}
    if "Propiedad" not in df.columns:
        return idx
    for i, v in enumerate(df["Propiedad"].tolist()):
        p = canon_prop(v, alias_prop)
        if p:
            idx[p] = i
    return idx


def _mapa_cortes(df: pd.DataFrame) -> Dict[str, str]:
    out: Dict[str, str] = {}
    META = {"propiedad", "unidad", "validacion", "validación", "validacion auto", "validación auto"}
    for c in df.columns:
        cc = canon_corte(str(c))
        if cc and str(c).strip().lower() not in META:
            out[cc] = str(c)
    return out


def _float_or_none(x: Any) -> Optional[float]:
    if x is None:
        return None
    try:
        s = str(x).strip()
        if s == "":
            return None
        return float(s.replace(",", "."))
    except Exception:
        return None


def _nombre_base_crudo(fname: str) -> str:
    base = re.sub(r"\.[^.]+$", "", fname)
    m = re.search(r"([A-Za-z]{3}-\d{4}-\d+)", base)
    if m:
        return m.group(1).upper()
    base = re.sub(r"^(?:ISA|RAMS)[_\-]+", "", base, flags=re.IGNORECASE)
    base = re.sub(r"([_\-])(ISA|RAMS)(?:([_\-]?v?\d{1,3})?)$", "", base, flags=re.IGNORECASE)
    base = re.sub(r"([_\-])(ISA|RAMS)(?:[_\-].*)?$", "", base, flags=re.IGNORECASE)
    return base.strip("_- ")


# ---------------------------------------------------------------------------
# 9. Emparejamiento de archivos
# ---------------------------------------------------------------------------

def pair_files(
    isa_names: List[str],
    rams_names: List[str],
) -> Tuple[Dict[str, Tuple[str, str]], List[str], List[str]]:
    isa_map: Dict[str, str] = {}
    for n in isa_names:
        base = _nombre_base_crudo(n)
        isa_map[base] = n

    rams_map: Dict[str, str] = {}
    for n in rams_names:
        base = _nombre_base_crudo(n)
        rams_map[base] = n

    paired: Dict[str, Tuple[str, str]] = {}
    for base, isa_fname in isa_map.items():
        if base in rams_map:
            paired[base] = (isa_fname, rams_map[base])

    unpaired_isa  = [isa_map[b] for b in isa_map if b not in rams_map]
    unpaired_rams = [rams_map[b] for b in rams_map if b not in isa_map]

    if unpaired_isa:
        logger.warning("ISA sin par RAMS: %s", unpaired_isa)
    if unpaired_rams:
        logger.warning("RAMS sin par ISA: %s", unpaired_rams)

    return paired, unpaired_isa, unpaired_rams


# ---------------------------------------------------------------------------
# 10. Cálculo de errores por crudo
# ---------------------------------------------------------------------------

def calcular_errores_crudo_df(
    df_isa: pd.DataFrame,
    df_rams: pd.DataFrame,
    umbrales: UmbralesDict,
    alias_prop: Dict[str, str],
    pct_ok_amarillo: float,
    pct_rojo_rojo: float,
    hoja_resumen: Dict[str, Dict[str, str]],
    crude_name: str,
    # Mantenidos en la firma por compatibilidad; ya no se usan
    tol: float = DEFAULT_TOL,
    tol_pesados: float = DEFAULT_TOL_PESADOS,
) -> Tuple[pd.DataFrame, List[str], List[str]]:
   """
    Calcula errores y semáforos para un par ISA/RAMS.
    Devuelve (df_out, columnas_cortes_visibles, orden_props_local).
    df_out tiene columnas: Propiedad | Semaforo | Corte_peor | Error_peor | Umbral_peor | [cortes...]
    """
    if "Propiedad" not in df_isa.columns:
        raise ValueError("El archivo ISA no tiene columna 'Propiedad'.")
    if "Propiedad" not in df_rams.columns:
        raise ValueError("El archivo RAMS no tiene columna 'Propiedad'.")

    cortes_isa = detectar_cortes_en_df(df_isa)
    if not cortes_isa:
        raise ValueError("El archivo ISA no contiene cortes reconocibles.")

    idx_rams = _indice_prop(df_rams, alias_prop)
    cortes_map_rams = _mapa_cortes(df_rams)

    columnas_cortes_visibles = [cname for (cname, _cc) in cortes_isa]
    df_out_cols = (
        ["Propiedad", "Semaforo", "Corte_peor", "Error_peor", "Umbral_peor"]
        + columnas_cortes_visibles
    )
    df_out = pd.DataFrame(columns=df_out_cols)
    orden_props_local: List[str] = []

    for _, row_isa in df_isa.iterrows():
        prop_raw = row_isa.get("Propiedad")
        prop_canon = canon_prop(prop_raw, alias_prop)
        if not prop_canon:
            continue

        orden_props_local.append(prop_canon)

        if prop_canon not in idx_rams:
            continue

        row_rams = df_rams.iloc[idx_rams[prop_canon]]

        errores_fila: Dict[str, float] = {}
        fila_out: Dict[str, Any] = {"Propiedad": str(prop_raw)}

        for (cname_isa, cc_isa) in cortes_isa:
            col_rams = cortes_map_rams.get(cc_isa)
            if col_rams is None:
                fila_out[cname_isa] = None
                continue
            isa_val = _float_or_none(row_isa.get(cname_isa))
            rams_val = _float_or_none(row_rams.get(col_rams))
            if isa_val is None or rams_val is None:
                err = None
            else:
                err = abs(isa_val - rams_val)
            errores_fila[cc_isa] = err
            fila_out[cname_isa] = err

        sem, _estados, corte_peor, error_peor, _ratio, umbral_peor = clasificar_propiedad(
            errores_fila,
            prop_canon,
            umbrales,
            tol=tol,
            pct_ok_amarillo=pct_ok_amarillo,
            pct_rojo_rojo=pct_rojo_rojo,
            tol_pesados=tol_pesados,
        )

        fila_out["Semaforo"] = sem
        fila_out["Corte_peor"] = corte_peor
        fila_out["Error_peor"] = error_peor
        fila_out["Umbral_peor"] = umbral_peor

        # Resumen global (propiedad × crudo)
        hoja_resumen.setdefault(prop_canon, {})[crude_name] = sem

        df_out.loc[len(df_out)] = fila_out

    return df_out, columnas_cortes_visibles, orden_props_local

# ---------------------------------------------------------------------------
# 12. Pipeline completo
# ---------------------------------------------------------------------------

def run_validation(
    isa_files: Dict[str, IO[bytes]],
    rams_files: Dict[str, IO[bytes]],
    matriz_file: IO[bytes],
    matriz_filename: str,
    pct_ok_amarillo: float = DEFAULT_PCT_OK_AMARILLO,
    pct_rojo_rojo: float = DEFAULT_PCT_ROJO_ROJO,
    sheet_hint: Optional[str] = None,
    # Mantenidos en la firma por compatibilidad con llamadas externas
    tol: float = DEFAULT_TOL,
    tol_pesados: float = DEFAULT_TOL_PESADOS,
) -> ValidationResult:
    """
    Pipeline completo: lee umbrales → empareja → calcula → construye ValidationResult.

    La clasificación por corte usa directamente los umbrales REPRO y ADMISIBLE
    de la matriz; los parámetros tol / tol_pesados se ignoran en la clasificación
    pero se siguen validando para mantener compatibilidad con la API anterior.
    """
    validate_params(tol, tol_pesados, pct_ok_amarillo, pct_rojo_rojo)

    alias_prop = crear_semantica_alias()
    df_matriz  = read_file_with_sheet(matriz_file, matriz_filename, sheet_hint)
    umbrales   = construir_umbrales(df_matriz, alias_prop)
    logger.info("Umbrales cargados: %d claves (repro + admisible)", len(umbrales))

    paired_map, unpaired_isa, unpaired_rams = pair_files(
        list(isa_files.keys()), list(rams_files.keys())
    )

    result = ValidationResult(unpaired_isa=unpaired_isa, unpaired_rams=unpaired_rams)
    resumen: Dict[str, Dict[str, str]] = {}
    orden_propiedades: List[str] = []

    for crude_name, (isa_fname, rams_fname) in sorted(paired_map.items()):
        try:
            logger.info("Procesando crudo: %s", crude_name)
            isa_files[isa_fname].seek(0)
            rams_files[rams_fname].seek(0)

            df_isa  = read_file(isa_files[isa_fname],  isa_fname)
            df_rams = read_file(rams_files[rams_fname], rams_fname)

            df_out, cortes_visibles, orden_local = calcular_errores_crudo_df(
                df_isa=df_isa,
                df_rams=df_rams,
                umbrales=umbrales,
                alias_prop=alias_prop,
                pct_ok_amarillo=pct_ok_amarillo,
                pct_rojo_rojo=pct_rojo_rojo,
                hoja_resumen=resumen,
                crude_name=crude_name,
            )

            result.paired_names.append(crude_name)
            result.crudo_dataframes[crude_name] = df_out
            result.cortes_visibles[crude_name]  = cortes_visibles

            if not orden_propiedades:
                orden_propiedades = orden_local

        except Exception as e:
            logger.error("Error procesando '%s': %s", crude_name, e)
            result.unpaired_isa.append(f"{isa_fname} [ERROR: {e}]")

    result.resumen_raw        = resumen
    result.orden_propiedades  = orden_propiedades
    result.pct_ok_amarillo    = pct_ok_amarillo
    result.pct_rojo_rojo      = pct_rojo_rojo
    result.summary = _build_summary_df(resumen, orden_propiedades, pct_ok_amarillo, pct_rojo_rojo)

    return result


# ---------------------------------------------------------------------------
# 13. DataFrame resumen
# ---------------------------------------------------------------------------

def _build_summary_df(
    resumen: Dict[str, Dict[str, str]],
    orden_propiedades: List[str],
    pct_ok_amarillo: float,
    pct_rojo_rojo: float,
) -> pd.DataFrame:
    todos_crudos    = sorted({c for m in resumen.values() for c in m.keys()})
    global_por_crudo = _sem_global_por_crudo(resumen, pct_ok_amarillo, pct_rojo_rojo)

    data = []
    if todos_crudos:
        row_global: Dict[str, Any] = {"Propiedad": "GLOBAL"}
        for cr in todos_crudos:
            row_global[cr] = global_por_crudo.get(cr, "")
        data.append(row_global)

    for prop in orden_propiedades:
        if prop in resumen:
            row: Dict[str, Any] = {"Propiedad": prop}
            for cr in todos_crudos:
                row[cr] = resumen[prop].get(cr, "")
            data.append(row)

    if data:
        return pd.DataFrame(data)
    return pd.DataFrame(columns=["Propiedad"])


# ---------------------------------------------------------------------------
# 14. Exportación a Excel
# ---------------------------------------------------------------------------

def add_conditional_formatting_text(ws, cell_range: str) -> None:
    fill_verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_amar  = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    fill_rojo  = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    fill_gris  = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

    dxf_v = DifferentialStyle(fill=fill_verde)
    dxf_a = DifferentialStyle(fill=fill_amar)
    dxf_r = DifferentialStyle(fill=fill_rojo)
    dxf_g = DifferentialStyle(fill=fill_gris)

    rule_v = Rule(type="containsText", operator="containsText", text="VERDE",    dxf=dxf_v)
    rule_a = Rule(type="containsText", operator="containsText", text="AMARILLO", dxf=dxf_a)
    rule_r = Rule(type="containsText", operator="containsText", text="ROJO",     dxf=dxf_r)
    rule_n = Rule(type="containsText", operator="containsText", text="NA",       dxf=dxf_g)

    ws.conditional_formatting.add(cell_range, rule_v)
    ws.conditional_formatting.add(cell_range, rule_a)
    ws.conditional_formatting.add(cell_range, rule_r)
    ws.conditional_formatting.add(cell_range, rule_n)


def _escribir_hoja_df(ws, df: pd.DataFrame) -> None:
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0) for cell in col
        )
        ws.column_dimensions[col[0].column_letter].width = min(max(10, max_len + 2), 45)


def build_excel(result: ValidationResult) -> bytes:
    wb  = Workbook()
    ws0 = wb.active
    ws0.title = "Resumen"

    df_res = result.summary
    _escribir_hoja_df(ws0, df_res)

    todos_crudos = [c for c in df_res.columns if c != "Propiedad"]
    if df_res.shape[0] > 0 and todos_crudos:
        start_row = 2
        end_row   = df_res.shape[0] + 1
        for j in range(2, 2 + len(todos_crudos)):
            col_letter = get_column_letter(j)
            add_conditional_formatting_text(ws0, f"{col_letter}{start_row}:{col_letter}{end_row}")

    for crude_name in result.paired_names:
        df_out = result.crudo_dataframes.get(crude_name, pd.DataFrame())
        hoja   = crude_name[:31]
        ws     = wb.create_sheet(title=hoja)
        _escribir_hoja_df(ws, df_out)
        if df_out.shape[0] > 0:
            add_conditional_formatting_text(ws, f"B2:B{df_out.shape[0] + 1}")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# Backwards-compat stubs
# ---------------------------------------------------------------------------

def canonize_name(name: str) -> str:
    return _nombre_base_crudo(name)


def validate_thresholds(config: "ThresholdConfig") -> None:
    if not (0 <= config.default_green < config.default_yellow):
        raise ValueError(
            f"Umbrales globales inválidos: verde={config.default_green}, "
            f"amarillo={config.default_yellow}. Se requiere 0 ≤ verde < amarillo."
        )
    for prop, (green, yellow) in config.thresholds.items():
        if not (0 <= green < yellow):
            raise ValueError(
                f"Umbrales inválidos para '{prop}': verde={green}, amarillo={yellow}."
            )
